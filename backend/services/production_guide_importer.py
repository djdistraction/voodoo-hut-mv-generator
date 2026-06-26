"""
Production Guide Importer
Parses production documents (XLSX, PDF, DOCX) into shot manifests & continuity bibles.

Supports the WOW OH! production format (shot sheet + continuity bible).
"""
import json
import re
from pathlib import Path
from typing import Optional
from database import (
    db_create_shot_manifest,
    db_update_series,
    db_get_series,
)

def parse_excel_shot_sheet(file_path: str, series_id: str = None, project_id: str = None) -> dict:
    """
    Parse an Excel shot sheet (like WOW_OH_Production_Shot_Sheet.xlsx).
    Returns {
        'shots': [list of shot manifests],
        'continuity_bible': {continuity rules},
        'metadata': {dashboard info}
    }
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl required: pip install openpyxl")

    workbook = openpyxl.load_workbook(file_path)

    result = {
        'shots': [],
        'continuity_bible': {},
        'metadata': {},
    }

    # Sheet 1: Dashboard (overview)
    if 'Dashboard' in workbook.sheetnames:
        result['metadata'] = _parse_dashboard(workbook['Dashboard'])

    # Sheet 2: Shot Sheet (main data)
    if 'Shot Sheet' in workbook.sheetnames:
        result['shots'] = _parse_shot_sheet(workbook['Shot Sheet'], project_id)

    # Sheet 3: Continuity Bible
    if 'Continuity Bible' in workbook.sheetnames:
        result['continuity_bible'] = _parse_continuity_bible(workbook['Continuity Bible'])

    # Sheet 4: Lookups (enums)
    if 'Lookups' in workbook.sheetnames:
        result['lookups'] = _parse_lookups(workbook['Lookups'])

    return result


def _parse_dashboard(sheet) -> dict:
    """Extract metadata from Dashboard sheet."""
    metadata = {}
    for row in sheet.iter_rows(min_row=1, max_row=20, values_only=True):
        if row[0] and row[1]:
            key = str(row[0]).lower().replace(" ", "_")
            metadata[key] = row[1]
    return metadata


def _parse_shot_sheet(sheet, project_id: str = None) -> list:
    """Extract shot manifests from Shot Sheet."""
    shots = []

    # Find header row (first row with actual column names)
    header_row = None
    for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True)):
        if row[0] and "Shot" in str(row[0]):
            header_row = i + 1
            headers = [str(h).lower().replace(" ", "_").replace("/", "_") if h else f"col_{j}"
                       for j, h in enumerate(row)]
            break

    if not header_row:
        raise ValueError("Could not find shot sheet header row")

    # Parse data rows
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        if not row[0]:  # skip empty rows
            continue

        shot_dict = dict(zip(headers, row))

        # Clean up None values
        shot_dict = {k: v for k, v in shot_dict.items() if v is not None}

        # Extract core fields
        shot_number = str(shot_dict.get('shot_', ''))
        timecode = str(shot_dict.get('timecode', ''))
        duration = str(shot_dict.get('duration', ''))
        section = shot_dict.get('section', '')
        audio_cue = shot_dict.get('audio_lyric', '')
        location = shot_dict.get('location', '')
        characters = shot_dict.get('characters', '')
        camera = shot_dict.get('camera', '')
        action = shot_dict.get('action', '')
        mood = shot_dict.get('mood', '')
        continuity = shot_dict.get('continuity_props', '')
        status = shot_dict.get('status', 'Not Started')
        priority = shot_dict.get('priority', '')
        notes = shot_dict.get('notes', '')

        # Parse timecode into start/end times
        start_time, end_time = _parse_timecode_range(timecode, duration)

        # Build character list
        char_list = [c.strip() for c in characters.split(',') if c.strip()] if isinstance(characters, str) else []

        # Build continuity rules list
        continuity_list = [c.strip() for c in continuity.split(';') if c.strip()] if isinstance(continuity, str) else []

        # Build negative constraints (inferred from notes/rules)
        negative_constraints = _infer_negative_constraints(shot_dict, audio_cue, characters)

        shot = {
            'shot_number': shot_number,
            'start_time': start_time,
            'end_time': end_time,
            'audio_cue': audio_cue,
            'location': location,
            'characters': char_list,
            'camera': camera,
            'action': action,
            'mood': mood,
            'continuity_rules': continuity_list,
            'negative_constraints': negative_constraints,
            'status': 'draft',  # always start as draft for review
            'section': section,
            'priority': priority,
            'original_status': status,
            'notes': notes,
        }

        shots.append(shot)

    return shots


def _parse_continuity_bible(sheet) -> dict:
    """Extract continuity bible from Continuity Bible sheet."""
    bible = {}

    for row in sheet.iter_rows(min_row=1, values_only=True):
        if row[0] and row[1]:
            key = str(row[0]).lower().replace(" ", "_")

            # Handle multi-line values or lists
            if isinstance(row[1], str) and '\n' in row[1]:
                value = [v.strip() for v in row[1].split('\n') if v.strip()]
            else:
                value = row[1]

            bible[key] = value

    return bible


def _parse_lookups(sheet) -> dict:
    """Extract lookup tables (enums) from Lookups sheet."""
    lookups = {}

    current_enum = None
    current_values = []

    for row in sheet.iter_rows(min_row=1, values_only=True):
        if not row[0]:
            continue

        # New enum section
        if row[0] and not row[1]:
            if current_enum and current_values:
                lookups[current_enum] = current_values
            current_enum = str(row[0]).lower().replace(" ", "_")
            current_values = []
        elif current_enum and row[0]:
            current_values.append(str(row[0]))

    if current_enum and current_values:
        lookups[current_enum] = current_values

    return lookups


def _parse_timecode_range(timecode: str, duration: str) -> tuple[str, str]:
    """Convert timecode + duration into start_time, end_time."""
    start = str(timecode).strip() if timecode else "0:00"
    dur = str(duration).strip() if duration else "0:00"

    try:
        # Try MM:SS format
        start_sec = _timecode_to_seconds(start)
        dur_sec = _timecode_to_seconds(dur)
        end_sec = start_sec + dur_sec
        return start, _seconds_to_timecode(end_sec)
    except:
        return start, ""


def _timecode_to_seconds(tc: str) -> float:
    """Convert MM:SS or MM:SS.sss to seconds."""
    parts = tc.split(':')
    minutes = int(parts[0]) if len(parts) > 0 else 0
    seconds = float(parts[1]) if len(parts) > 1 else 0
    return minutes * 60 + seconds


def _seconds_to_timecode(sec: float) -> str:
    """Convert seconds to MM:SS format."""
    minutes = int(sec // 60)
    secs = sec % 60
    return f"{minutes}:{secs:05.2f}"


def _infer_negative_constraints(shot_dict: dict, audio_cue: str, characters: str) -> list:
    """Infer negative constraints from shot data and notes."""
    constraints = []

    notes = shot_dict.get('notes', '').lower() if isinstance(shot_dict.get('notes'), str) else ""

    # Common negative constraints by context
    if 'stress' in str(audio_cue).lower() or 'stress' in notes:
        constraints.extend([
            "no calming colors",
            "no peaceful expressions",
            "no static poses",
        ])

    if 'office' in str(shot_dict.get('location', '')).lower():
        constraints.extend([
            "no outdoor scenery",
            "no vibrant neon colors",
        ])

    if 'club' in str(shot_dict.get('location', '')).lower():
        constraints.extend([
            "no daylight",
            "no corporate office elements",
        ])

    # Avoid text in image generation
    if 'text' not in notes and 'logo' not in notes:
        constraints.append("no text or logos")

    return list(set(constraints))  # deduplicate


def import_wow_oh_series(series_id: str, excel_path: str) -> dict:
    """
    Import WOW OH! production guide as a complete series + set of project manifests.
    Creates series record with continuity bible, returns shot manifests ready for project binding.
    """
    parsed = parse_excel_shot_sheet(excel_path, series_id=series_id)

    # Update series with continuity bible
    if parsed['continuity_bible']:
        db_update_series(series_id, continuity_bible=parsed['continuity_bible'])

    return {
        'shots': parsed['shots'],
        'continuity_bible': parsed['continuity_bible'],
        'metadata': parsed['metadata'],
        'lookups': parsed.get('lookups', {}),
    }


def create_project_shot_manifests(project_id: str, shots_data: list) -> list[str]:
    """
    Create shot manifest records in database for a project.
    Returns list of manifest IDs.
    """
    manifest_ids = []

    for shot in shots_data:
        manifest_id = db_create_shot_manifest(
            project_id=project_id,
            shot_number=shot['shot_number'],
            start_time=shot['start_time'],
            end_time=shot['end_time'],
            audio_cue=shot.get('audio_cue', ''),
            location=shot.get('location', ''),
            characters=shot.get('characters', []),
            camera=shot.get('camera', ''),
            action=shot.get('action', ''),
            mood=shot.get('mood', ''),
            continuity_rules=shot.get('continuity_rules', []),
            negative_constraints=shot.get('negative_constraints', []),
        )
        manifest_ids.append(manifest_id)

    return manifest_ids
